from flask import Blueprint, request, jsonify, send_file
from src.routes.auth import token_required
from src.utils.subscription_utils import verificar_limite_plano
from src.models.financeiro import (
    CategoriaFinanceira, TransacaoFinanceira, MetaFinanceira, 
    ContaFinanceira, LembreteFinanceiro, OrcamentoMensal
)
from src.models.user import db
from datetime import datetime, date, timedelta
from dateutil.relativedelta import relativedelta
import calendar
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import PieChart, BarChart, Reference
import openai
import os

financeiro_bp = Blueprint('financeiro', __name__)

# ==================== TRANSAÇÕES ====================

@financeiro_bp.route('/transacoes', methods=['GET'])
@token_required
def listar_transacoes(current_user):
    """Lista transações financeiras com filtros"""
    try:
        # Parâmetros de filtro
        mes = request.args.get('mes', type=int)
        ano = request.args.get('ano', type=int)
        categoria_id = request.args.get('categoria_id', type=int)
        tipo = request.args.get('tipo')  # 'receita' ou 'despesa'
        limit = request.args.get('limit', 50, type=int)
        
        query = TransacaoFinanceira.query.filter_by(user_id=current_user.id)
        
        # Aplicar filtros
        if mes and ano:
            inicio_mes = date(ano, mes, 1)
            fim_mes = date(ano, mes, calendar.monthrange(ano, mes)[1])
            query = query.filter(TransacaoFinanceira.data_transacao.between(inicio_mes, fim_mes))
        
        if categoria_id:
            query = query.filter_by(categoria_id=categoria_id)
        
        if tipo:
            query = query.filter_by(tipo=tipo)
        
        transacoes = query.order_by(TransacaoFinanceira.data_transacao.desc()).limit(limit).all()
        
        return jsonify({
            'transacoes': [transacao.to_dict() for transacao in transacoes],
            'total': query.count()
        }), 200
        
    except Exception as e:
        return jsonify({'erro': f'Erro interno: {str(e)}'}), 500

@financeiro_bp.route('/transacoes', methods=['POST'])
@token_required
@verificar_limite_plano(incrementar_uso='transacoes')
def criar_transacao(current_user):
    """Cria nova transação financeira"""
    try:
        data = request.get_json()
        
        # Validações
        campos_obrigatorios = ['descricao', 'valor', 'tipo', 'categoria_id']
        for campo in campos_obrigatorios:
            if not data.get(campo):
                return jsonify({'erro': f'Campo {campo} é obrigatório'}), 400
        
        # Verifica se categoria existe e pertence ao usuário
        categoria = CategoriaFinanceira.query.filter_by(
            id=data['categoria_id'], 
            user_id=current_user.id
        ).first()
        
        if not categoria:
            return jsonify({'erro': 'Categoria não encontrada'}), 404
        
        # Cria transação
        transacao = TransacaoFinanceira(
            user_id=current_user.id,
            categoria_id=data['categoria_id'],
            descricao=data['descricao'],
            valor=data['valor'],
            tipo=data['tipo'],
            data_transacao=datetime.strptime(data.get('data_transacao', date.today().isoformat()), '%Y-%m-%d').date(),
            observacoes=data.get('observacoes'),
            anexo_url=data.get('anexo_url'),
            recorrente=data.get('recorrente', False),
            frequencia_recorrencia=data.get('frequencia_recorrencia')
        )
        
        # Define tags se fornecidas
        if data.get('tags'):
            transacao.set_tags(data['tags'])
        
        # Calcula próxima ocorrência se recorrente
        if transacao.recorrente and transacao.frequencia_recorrencia:
            if transacao.frequencia_recorrencia == 'mensal':
                transacao.proxima_ocorrencia = transacao.data_transacao + relativedelta(months=1)
            elif transacao.frequencia_recorrencia == 'semanal':
                transacao.proxima_ocorrencia = transacao.data_transacao + timedelta(weeks=1)
            elif transacao.frequencia_recorrencia == 'anual':
                transacao.proxima_ocorrencia = transacao.data_transacao + relativedelta(years=1)
        
        db.session.add(transacao)
        db.session.commit()
        
        # Atualiza orçamento mensal se existir
        atualizar_orcamento_mensal(current_user.id, transacao)
        
        return jsonify({
            'message': 'Transação criada com sucesso',
            'transacao': transacao.to_dict()
        }), 201
        
    except Exception as e:
        return jsonify({'erro': f'Erro interno: {str(e)}'}), 500

@financeiro_bp.route('/transacoes/<int:transacao_id>', methods=['PUT'])
@token_required
def atualizar_transacao(current_user, transacao_id):
    """Atualiza transação existente"""
    try:
        transacao = TransacaoFinanceira.query.filter_by(
            id=transacao_id, 
            user_id=current_user.id
        ).first()
        
        if not transacao:
            return jsonify({'erro': 'Transação não encontrada'}), 404
        
        data = request.get_json()
        
        # Atualiza campos
        if 'descricao' in data:
            transacao.descricao = data['descricao']
        if 'valor' in data:
            transacao.valor = data['valor']
        if 'tipo' in data:
            transacao.tipo = data['tipo']
        if 'categoria_id' in data:
            transacao.categoria_id = data['categoria_id']
        if 'data_transacao' in data:
            transacao.data_transacao = datetime.strptime(data['data_transacao'], '%Y-%m-%d').date()
        if 'observacoes' in data:
            transacao.observacoes = data['observacoes']
        if 'tags' in data:
            transacao.set_tags(data['tags'])
        
        db.session.commit()
        
        return jsonify({
            'message': 'Transação atualizada com sucesso',
            'transacao': transacao.to_dict()
        }), 200
        
    except Exception as e:
        return jsonify({'erro': f'Erro interno: {str(e)}'}), 500

@financeiro_bp.route('/transacoes/<int:transacao_id>', methods=['DELETE'])
@token_required
def excluir_transacao(current_user, transacao_id):
    """Exclui transação"""
    try:
        transacao = TransacaoFinanceira.query.filter_by(
            id=transacao_id, 
            user_id=current_user.id
        ).first()
        
        if not transacao:
            return jsonify({'erro': 'Transação não encontrada'}), 404
        
        db.session.delete(transacao)
        db.session.commit()
        
        return jsonify({'message': 'Transação excluída com sucesso'}), 200
        
    except Exception as e:
        return jsonify({'erro': f'Erro interno: {str(e)}'}), 500

# ==================== CATEGORIAS ====================

@financeiro_bp.route('/categorias', methods=['GET'])
@token_required
def listar_categorias(current_user):
    """Lista categorias financeiras"""
    try:
        tipo = request.args.get('tipo')  # 'receita' ou 'despesa'
        
        query = CategoriaFinanceira.query.filter_by(user_id=current_user.id, ativa=True)
        
        if tipo:
            query = query.filter_by(tipo=tipo)
        
        categorias = query.order_by(CategoriaFinanceira.nome).all()
        
        return jsonify({
            'categorias': [categoria.to_dict() for categoria in categorias]
        }), 200
        
    except Exception as e:
        return jsonify({'erro': f'Erro interno: {str(e)}'}), 500

@financeiro_bp.route('/categorias', methods=['POST'])
@token_required
def criar_categoria(current_user):
    """Cria nova categoria financeira"""
    try:
        data = request.get_json()
        
        if not data.get('nome') or not data.get('tipo'):
            return jsonify({'erro': 'Nome e tipo são obrigatórios'}), 400
        
        categoria = CategoriaFinanceira(
            user_id=current_user.id,
            nome=data['nome'],
            tipo=data['tipo'],
            cor=data.get('cor', '#3b82f6'),
            icone=data.get('icone', '💰')
        )
        
        db.session.add(categoria)
        db.session.commit()
        
        return jsonify({
            'message': 'Categoria criada com sucesso',
            'categoria': categoria.to_dict()
        }), 201
        
    except Exception as e:
        return jsonify({'erro': f'Erro interno: {str(e)}'}), 500

# ==================== DASHBOARD FINANCEIRO ====================

@financeiro_bp.route('/dashboard', methods=['GET'])
@token_required
def dashboard_financeiro(current_user):
    """Retorna dados do dashboard financeiro"""
    try:
        mes_atual = date.today().month
        ano_atual = date.today().year
        
        # Receitas e despesas do mês atual
        inicio_mes = date(ano_atual, mes_atual, 1)
        fim_mes = date(ano_atual, mes_atual, calendar.monthrange(ano_atual, mes_atual)[1])
        
        transacoes_mes = TransacaoFinanceira.query.filter_by(user_id=current_user.id)\
                                                .filter(TransacaoFinanceira.data_transacao.between(inicio_mes, fim_mes))\
                                                .all()
        
        receitas_mes = sum(float(t.valor) for t in transacoes_mes if t.tipo == 'receita')
        despesas_mes = sum(float(t.valor) for t in transacoes_mes if t.tipo == 'despesa')
        saldo_mes = receitas_mes - despesas_mes
        
        # Saldo total das contas
        contas = ContaFinanceira.query.filter_by(user_id=current_user.id, ativa=True, incluir_no_total=True).all()
        saldo_total = sum(float(conta.saldo_atual) for conta in contas)
        
        # Metas ativas
        metas_ativas = MetaFinanceira.query.filter_by(user_id=current_user.id, status='ativa').count()
        
        # Lembretes pendentes
        lembretes_pendentes = LembreteFinanceiro.query.filter_by(
            user_id=current_user.id, 
            status='pendente'
        ).filter(LembreteFinanceiro.data_vencimento >= date.today()).count()
        
        # Transações por categoria (últimos 30 dias)
        data_limite = date.today() - timedelta(days=30)
        transacoes_recentes = TransacaoFinanceira.query.filter_by(user_id=current_user.id)\
                                                     .filter(TransacaoFinanceira.data_transacao >= data_limite)\
                                                     .all()
        
        gastos_por_categoria = {}
        for transacao in transacoes_recentes:
            if transacao.tipo == 'despesa' and transacao.categoria:
                nome_categoria = transacao.categoria.nome
                gastos_por_categoria[nome_categoria] = gastos_por_categoria.get(nome_categoria, 0) + float(transacao.valor)
        
        # Top 5 categorias de gastos
        top_categorias = sorted(gastos_por_categoria.items(), key=lambda x: x[1], reverse=True)[:5]
        
        # Evolução mensal (últimos 6 meses)
        evolucao_mensal = []
        for i in range(6):
            data_ref = date.today() - relativedelta(months=i)
            inicio = date(data_ref.year, data_ref.month, 1)
            fim = date(data_ref.year, data_ref.month, calendar.monthrange(data_ref.year, data_ref.month)[1])
            
            transacoes_periodo = TransacaoFinanceira.query.filter_by(user_id=current_user.id)\
                                                        .filter(TransacaoFinanceira.data_transacao.between(inicio, fim))\
                                                        .all()
            
            receitas = sum(float(t.valor) for t in transacoes_periodo if t.tipo == 'receita')
            despesas = sum(float(t.valor) for t in transacoes_periodo if t.tipo == 'despesa')
            
            evolucao_mensal.append({
                'mes': data_ref.strftime('%m/%Y'),
                'receitas': receitas,
                'despesas': despesas,
                'saldo': receitas - despesas
            })
        
        evolucao_mensal.reverse()
        
        return jsonify({
            'resumo': {
                'receitas_mes': receitas_mes,
                'despesas_mes': despesas_mes,
                'saldo_mes': saldo_mes,
                'saldo_total': saldo_total,
                'metas_ativas': metas_ativas,
                'lembretes_pendentes': lembretes_pendentes
            },
            'gastos_por_categoria': [
                {'categoria': cat, 'valor': valor} 
                for cat, valor in top_categorias
            ],
            'evolucao_mensal': evolucao_mensal,
            'contas': [conta.to_dict() for conta in contas]
        }), 200
        
    except Exception as e:
        return jsonify({'erro': f'Erro interno: {str(e)}'}), 500

# ==================== LEMBRETES FINANCEIROS ====================

@financeiro_bp.route('/lembretes', methods=['GET'])
@token_required
def listar_lembretes(current_user):
    """Lista lembretes financeiros"""
    try:
        status = request.args.get('status')
        proximos_dias = request.args.get('proximos_dias', type=int)
        
        query = LembreteFinanceiro.query.filter_by(user_id=current_user.id)
        
        if status:
            query = query.filter_by(status=status)
        
        if proximos_dias:
            data_limite = date.today() + timedelta(days=proximos_dias)
            query = query.filter(LembreteFinanceiro.data_vencimento <= data_limite)
        
        lembretes = query.order_by(LembreteFinanceiro.data_vencimento).all()
        
        return jsonify({
            'lembretes': [lembrete.to_dict() for lembrete in lembretes]
        }), 200
        
    except Exception as e:
        return jsonify({'erro': f'Erro interno: {str(e)}'}), 500

@financeiro_bp.route('/lembretes', methods=['POST'])
@token_required
def criar_lembrete(current_user):
    """Cria novo lembrete financeiro"""
    try:
        data = request.get_json()
        
        campos_obrigatorios = ['titulo', 'tipo_lembrete', 'data_vencimento']
        for campo in campos_obrigatorios:
            if not data.get(campo):
                return jsonify({'erro': f'Campo {campo} é obrigatório'}), 400
        
        lembrete = LembreteFinanceiro(
            user_id=current_user.id,
            titulo=data['titulo'],
            descricao=data.get('descricao'),
            tipo_lembrete=data['tipo_lembrete'],
            valor=data.get('valor'),
            data_vencimento=datetime.strptime(data['data_vencimento'], '%Y-%m-%d').date(),
            recorrente=data.get('recorrente', False),
            frequencia=data.get('frequencia'),
            dias_antecedencia=data.get('dias_antecedencia', 3),
            prioridade=data.get('prioridade', 'media')
        )
        
        db.session.add(lembrete)
        db.session.commit()
        
        return jsonify({
            'message': 'Lembrete criado com sucesso',
            'lembrete': lembrete.to_dict()
        }), 201
        
    except Exception as e:
        return jsonify({'erro': f'Erro interno: {str(e)}'}), 500

@financeiro_bp.route('/lembretes/<int:lembrete_id>/concluir', methods=['PUT'])
@token_required
def concluir_lembrete(current_user, lembrete_id):
    """Marca lembrete como concluído e opcionalmente cria transação"""
    try:
        lembrete = LembreteFinanceiro.query.filter_by(
            id=lembrete_id, 
            user_id=current_user.id
        ).first()
        
        if not lembrete:
            return jsonify({'erro': 'Lembrete não encontrado'}), 404
        
        data = request.get_json()
        criar_transacao_flag = data.get('criar_transacao', False)
        
        # Marca como concluído
        lembrete.status = 'concluido'
        
        # Cria transação se solicitado
        if criar_transacao_flag and lembrete.valor:
            # Busca categoria padrão ou usa a primeira disponível
            categoria = CategoriaFinanceira.query.filter_by(
                user_id=current_user.id,
                tipo='despesa' if lembrete.tipo_lembrete == 'pagamento' else 'receita',
                ativa=True
            ).first()
            
            if categoria:
                transacao = TransacaoFinanceira(
                    user_id=current_user.id,
                    categoria_id=categoria.id,
                    descricao=lembrete.titulo,
                    valor=lembrete.valor,
                    tipo='despesa' if lembrete.tipo_lembrete == 'pagamento' else 'receita',
                    data_transacao=date.today(),
                    observacoes=f'Criado automaticamente do lembrete: {lembrete.titulo}'
                )
                
                db.session.add(transacao)
                lembrete.transacao_id = transacao.id
        
        # Cria próximo lembrete se recorrente
        if lembrete.recorrente and lembrete.frequencia:
            proximo_lembrete = LembreteFinanceiro(
                user_id=current_user.id,
                titulo=lembrete.titulo,
                descricao=lembrete.descricao,
                tipo_lembrete=lembrete.tipo_lembrete,
                valor=lembrete.valor,
                recorrente=True,
                frequencia=lembrete.frequencia,
                dias_antecedencia=lembrete.dias_antecedencia,
                prioridade=lembrete.prioridade
            )
            
            # Calcula próxima data
            if lembrete.frequencia == 'mensal':
                proximo_lembrete.data_vencimento = lembrete.data_vencimento + relativedelta(months=1)
            elif lembrete.frequencia == 'anual':
                proximo_lembrete.data_vencimento = lembrete.data_vencimento + relativedelta(years=1)
            elif lembrete.frequencia == 'semanal':
                proximo_lembrete.data_vencimento = lembrete.data_vencimento + timedelta(weeks=1)
            
            db.session.add(proximo_lembrete)
        
        db.session.commit()
        
        return jsonify({
            'message': 'Lembrete concluído com sucesso',
            'lembrete': lembrete.to_dict()
        }), 200
        
    except Exception as e:
        return jsonify({'erro': f'Erro interno: {str(e)}'}), 500

# ==================== RELATÓRIOS E PLANILHAS ====================

@financeiro_bp.route('/relatorio/excel', methods=['POST'])
@token_required
def gerar_relatorio_excel(current_user):
    """Gera relatório financeiro em Excel"""
    try:
        data = request.get_json()
        mes = data.get('mes', date.today().month)
        ano = data.get('ano', date.today().year)
        
        # Cria workbook
        wb = openpyxl.Workbook()
        
        # Remove sheet padrão
        wb.remove(wb.active)
        
        # ===== ABA RESUMO =====
        ws_resumo = wb.create_sheet("Resumo")
        
        # Título
        ws_resumo['A1'] = f'Relatório Financeiro - {mes:02d}/{ano}'
        ws_resumo['A1'].font = Font(size=16, bold=True)
        ws_resumo.merge_cells('A1:D1')
        
        # Busca dados do mês
        inicio_mes = date(ano, mes, 1)
        fim_mes = date(ano, mes, calendar.monthrange(ano, mes)[1])
        
        transacoes_mes = TransacaoFinanceira.query.filter_by(user_id=current_user.id)\
                                                .filter(TransacaoFinanceira.data_transacao.between(inicio_mes, fim_mes))\
                                                .all()
        
        receitas_total = sum(float(t.valor) for t in transacoes_mes if t.tipo == 'receita')
        despesas_total = sum(float(t.valor) for t in transacoes_mes if t.tipo == 'despesa')
        saldo = receitas_total - despesas_total
        
        # Resumo financeiro
        ws_resumo['A3'] = 'RESUMO FINANCEIRO'
        ws_resumo['A3'].font = Font(bold=True)
        
        ws_resumo['A4'] = 'Receitas:'
        ws_resumo['B4'] = f'R$ {receitas_total:,.2f}'
        ws_resumo['B4'].number_format = 'R$ #,##0.00'
        
        ws_resumo['A5'] = 'Despesas:'
        ws_resumo['B5'] = f'R$ {despesas_total:,.2f}'
        ws_resumo['B5'].number_format = 'R$ #,##0.00'
        
        ws_resumo['A6'] = 'Saldo:'
        ws_resumo['B6'] = f'R$ {saldo:,.2f}'
        ws_resumo['B6'].number_format = 'R$ #,##0.00'
        ws_resumo['B6'].font = Font(bold=True)
        if saldo < 0:
            ws_resumo['B6'].font = Font(bold=True, color='FF0000')
        else:
            ws_resumo['B6'].font = Font(bold=True, color='008000')
        
        # ===== ABA TRANSAÇÕES =====
        ws_transacoes = wb.create_sheet("Transações")
        
        # Cabeçalhos
        headers = ['Data', 'Descrição', 'Categoria', 'Tipo', 'Valor', 'Observações']
        for col, header in enumerate(headers, 1):
            cell = ws_transacoes.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
        
        # Dados das transações
        for row, transacao in enumerate(transacoes_mes, 2):
            ws_transacoes.cell(row=row, column=1, value=transacao.data_transacao)
            ws_transacoes.cell(row=row, column=2, value=transacao.descricao)
            ws_transacoes.cell(row=row, column=3, value=transacao.categoria.nome if transacao.categoria else '')
            ws_transacoes.cell(row=row, column=4, value=transacao.tipo.title())
            
            valor_cell = ws_transacoes.cell(row=row, column=5, value=float(transacao.valor))
            valor_cell.number_format = 'R$ #,##0.00'
            if transacao.tipo == 'despesa':
                valor_cell.font = Font(color='FF0000')
            else:
                valor_cell.font = Font(color='008000')
            
            ws_transacoes.cell(row=row, column=6, value=transacao.observacoes or '')
        
        # Ajusta largura das colunas
        for column in ws_transacoes.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_transacoes.column_dimensions[column_letter].width = adjusted_width
        
        # ===== ABA CATEGORIAS =====
        ws_categorias = wb.create_sheet("Por Categoria")
        
        # Agrupa por categoria
        gastos_categoria = {}
        receitas_categoria = {}
        
        for transacao in transacoes_mes:
            categoria_nome = transacao.categoria.nome if transacao.categoria else 'Sem categoria'
            valor = float(transacao.valor)
            
            if transacao.tipo == 'despesa':
                gastos_categoria[categoria_nome] = gastos_categoria.get(categoria_nome, 0) + valor
            else:
                receitas_categoria[categoria_nome] = receitas_categoria.get(categoria_nome, 0) + valor
        
        # Cabeçalhos
        ws_categorias['A1'] = 'DESPESAS POR CATEGORIA'
        ws_categorias['A1'].font = Font(bold=True)
        ws_categorias['A2'] = 'Categoria'
        ws_categorias['B2'] = 'Valor'
        
        row = 3
        for categoria, valor in sorted(gastos_categoria.items(), key=lambda x: x[1], reverse=True):
            ws_categorias.cell(row=row, column=1, value=categoria)
            valor_cell = ws_categorias.cell(row=row, column=2, value=valor)
            valor_cell.number_format = 'R$ #,##0.00'
            row += 1
        
        # Receitas por categoria
        ws_categorias[f'A{row+1}'] = 'RECEITAS POR CATEGORIA'
        ws_categorias[f'A{row+1}'].font = Font(bold=True)
        row += 2
        ws_categorias[f'A{row}'] = 'Categoria'
        ws_categorias[f'B{row}'] = 'Valor'
        row += 1
        
        for categoria, valor in sorted(receitas_categoria.items(), key=lambda x: x[1], reverse=True):
            ws_categorias.cell(row=row, column=1, value=categoria)
            valor_cell = ws_categorias.cell(row=row, column=2, value=valor)
            valor_cell.number_format = 'R$ #,##0.00'
            row += 1
        
        # Salva em buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return send_file(
            buffer,
            as_attachment=True,
            download_name=f'relatorio_financeiro_{mes:02d}_{ano}.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        return jsonify({'erro': f'Erro interno: {str(e)}'}), 500

# ==================== RECOMENDAÇÕES FINANCEIRAS ====================

@financeiro_bp.route('/recomendacoes', methods=['GET'])
@token_required
def obter_recomendacoes(current_user):
    """Gera recomendações financeiras personalizadas"""
    try:
        # Analisa dados financeiros do usuário
        analise = analisar_situacao_financeira(current_user.id)
        
        # Gera recomendações usando IA
        recomendacoes = gerar_recomendacoes_ia(analise)
        
        return jsonify({
            'analise': analise,
            'recomendacoes': recomendacoes,
            'gerado_em': datetime.now().isoformat()
        }), 200
        
    except Exception as e:
        return jsonify({'erro': f'Erro interno: {str(e)}'}), 500

def analisar_situacao_financeira(user_id):
    """Analisa situação financeira do usuário"""
    try:
        # Dados dos últimos 3 meses
        data_limite = date.today() - relativedelta(months=3)
        
        transacoes = TransacaoFinanceira.query.filter_by(user_id=user_id)\
                                            .filter(TransacaoFinanceira.data_transacao >= data_limite)\
                                            .all()
        
        # Cálculos básicos
        receitas_total = sum(float(t.valor) for t in transacoes if t.tipo == 'receita')
        despesas_total = sum(float(t.valor) for t in transacoes if t.tipo == 'despesa')
        saldo_periodo = receitas_total - despesas_total
        
        # Média mensal
        receita_mensal = receitas_total / 3
        despesa_mensal = despesas_total / 3
        
        # Análise por categoria
        gastos_categoria = {}
        for transacao in transacoes:
            if transacao.tipo == 'despesa' and transacao.categoria:
                categoria = transacao.categoria.nome
                gastos_categoria[categoria] = gastos_categoria.get(categoria, 0) + float(transacao.valor)
        
        # Categoria que mais gasta
        maior_gasto = max(gastos_categoria.items(), key=lambda x: x[1]) if gastos_categoria else ('N/A', 0)
        
        # Taxa de poupança
        taxa_poupanca = (saldo_periodo / receitas_total * 100) if receitas_total > 0 else 0
        
        # Contas e saldo total
        contas = ContaFinanceira.query.filter_by(user_id=user_id, ativa=True).all()
        saldo_total = sum(float(conta.saldo_atual) for conta in contas)
        
        # Metas
        metas = MetaFinanceira.query.filter_by(user_id=user_id, status='ativa').all()
        
        return {
            'periodo_analise': '3 meses',
            'receitas_total': receitas_total,
            'despesas_total': despesas_total,
            'saldo_periodo': saldo_periodo,
            'receita_mensal_media': receita_mensal,
            'despesa_mensal_media': despesa_mensal,
            'taxa_poupanca': taxa_poupanca,
            'saldo_atual': saldo_total,
            'maior_categoria_gasto': {
                'categoria': maior_gasto[0],
                'valor': maior_gasto[1]
            },
            'total_metas_ativas': len(metas),
            'gastos_por_categoria': gastos_categoria
        }
        
    except Exception as e:
        return {
            'erro': f'Erro na análise: {str(e)}',
            'periodo_analise': '3 meses',
            'receitas_total': 0,
            'despesas_total': 0
        }

def gerar_recomendacoes_ia(analise):
    """Gera recomendações usando IA"""
    try:
        prompt = f"""
        Baseado na análise financeira abaixo, gere 5 recomendações práticas e personalizadas:
        
        Situação Financeira:
        - Receita mensal média: R$ {analise['receita_mensal_media']:.2f}
        - Despesa mensal média: R$ {analise['despesa_mensal_media']:.2f}
        - Taxa de poupança: {analise['taxa_poupanca']:.1f}%
        - Saldo atual: R$ {analise['saldo_atual']:.2f}
        - Maior categoria de gasto: {analise['maior_categoria_gasto']['categoria']} (R$ {analise['maior_categoria_gasto']['valor']:.2f})
        - Metas ativas: {analise['total_metas_ativas']}
        
        Forneça recomendações específicas, práticas e acionáveis em português brasileiro.
        Cada recomendação deve ter um título e uma descrição de 2-3 linhas.
        """
        
        if os.getenv('OPENAI_API_KEY'):
            client = openai.OpenAI()
            response = client.chat.completions.create(
                model="gpt-3.5-turbo",
                messages=[{"role": "user", "content": prompt}],
                max_tokens=500,
                temperature=0.7
            )
            
            # Processa resposta da IA
            texto_recomendacoes = response.choices[0].message.content
            
            # Converte em lista estruturada (simplificado)
            recomendacoes = []
            linhas = texto_recomendacoes.split('\n')
            
            titulo_atual = ""
            descricao_atual = ""
            
            for linha in linhas:
                linha = linha.strip()
                if linha and (linha.startswith('1.') or linha.startswith('2.') or 
                            linha.startswith('3.') or linha.startswith('4.') or 
                            linha.startswith('5.') or '**' in linha):
                    if titulo_atual and descricao_atual:
                        recomendacoes.append({
                            'titulo': titulo_atual,
                            'descricao': descricao_atual.strip(),
                            'prioridade': 'alta' if len(recomendacoes) < 2 else 'media'
                        })
                    titulo_atual = linha.replace('**', '').replace('1.', '').replace('2.', '').replace('3.', '').replace('4.', '').replace('5.', '').strip()
                    descricao_atual = ""
                elif linha and titulo_atual:
                    descricao_atual += linha + " "
            
            # Adiciona última recomendação
            if titulo_atual and descricao_atual:
                recomendacoes.append({
                    'titulo': titulo_atual,
                    'descricao': descricao_atual.strip(),
                    'prioridade': 'media'
                })
            
            return recomendacoes[:5]  # Limita a 5 recomendações
            
        else:
            # Recomendações básicas sem IA
            return gerar_recomendacoes_basicas(analise)
            
    except Exception as e:
        return gerar_recomendacoes_basicas(analise)

def gerar_recomendacoes_basicas(analise):
    """Gera recomendações básicas sem IA"""
    recomendacoes = []
    
    # Recomendação baseada na taxa de poupança
    if analise['taxa_poupanca'] < 10:
        recomendacoes.append({
            'titulo': 'Aumente sua taxa de poupança',
            'descricao': f'Sua taxa de poupança atual é de {analise["taxa_poupanca"]:.1f}%. Tente economizar pelo menos 10-20% da sua renda mensal.',
            'prioridade': 'alta'
        })
    
    # Recomendação baseada no maior gasto
    if analise['maior_categoria_gasto']['valor'] > 0:
        recomendacoes.append({
            'titulo': f'Revise gastos com {analise["maior_categoria_gasto"]["categoria"]}',
            'descricao': f'Esta é sua maior categoria de gastos (R$ {analise["maior_categoria_gasto"]["valor"]:.2f}). Analise se há oportunidades de redução.',
            'prioridade': 'media'
        })
    
    # Recomendação sobre metas
    if analise['total_metas_ativas'] == 0:
        recomendacoes.append({
            'titulo': 'Defina metas financeiras',
            'descricao': 'Estabeleça metas claras como reserva de emergência, aposentadoria ou grandes compras para manter o foco.',
            'prioridade': 'media'
        })
    
    # Recomendação sobre reserva de emergência
    reserva_ideal = analise['despesa_mensal_media'] * 6
    if analise['saldo_atual'] < reserva_ideal:
        recomendacoes.append({
            'titulo': 'Construa sua reserva de emergência',
            'descricao': f'Recomenda-se ter 6 meses de gastos guardados (R$ {reserva_ideal:.2f}). Seu saldo atual: R$ {analise["saldo_atual"]:.2f}.',
            'prioridade': 'alta'
        })
    
    # Recomendação geral
    recomendacoes.append({
        'titulo': 'Monitore regularmente suas finanças',
        'descricao': 'Revise seus gastos semanalmente e ajuste seu orçamento conforme necessário para manter o controle financeiro.',
        'prioridade': 'baixa'
    })
    
    return recomendacoes[:5]

def atualizar_orcamento_mensal(user_id, transacao):
    """Atualiza orçamento mensal após nova transação"""
    try:
        if transacao.tipo != 'despesa':
            return
        
        orcamento = OrcamentoMensal.query.filter_by(
            user_id=user_id,
            categoria_id=transacao.categoria_id,
            mes=transacao.data_transacao.month,
            ano=transacao.data_transacao.year
        ).first()
        
        if orcamento:
            orcamento.valor_gasto += transacao.valor
            db.session.commit()
            
    except Exception as e:
        print(f"Erro ao atualizar orçamento: {e}")

# ==================== PROCESSAMENTO DE COMANDOS DE VOZ ====================

@financeiro_bp.route('/processar-comando', methods=['POST'])
@token_required
def processar_comando_financeiro(current_user):
    """Processa comando de voz relacionado a finanças"""
    try:
        data = request.get_json()
        comando = data.get('comando', '').lower()
        
        resposta = ""
        acao_executada = None
        dados_criados = None
        
        # Detecta tipo de comando
        if any(palavra in comando for palavra in ['gastei', 'paguei', 'comprei', 'despesa']):
            # Comando de despesa
            resultado = processar_comando_despesa(comando, current_user.id)
            resposta = resultado['resposta']
            acao_executada = 'criar_despesa'
            dados_criados = resultado.get('transacao')
            
        elif any(palavra in comando for palavra in ['recebi', 'ganhei', 'receita', 'salário']):
            # Comando de receita
            resultado = processar_comando_receita(comando, current_user.id)
            resposta = resultado['resposta']
            acao_executada = 'criar_receita'
            dados_criados = resultado.get('transacao')
            
        elif any(palavra in comando for palavra in ['lembrar', 'lembrete', 'pagar', 'vencimento']):
            # Comando de lembrete
            resultado = processar_comando_lembrete(comando, current_user.id)
            resposta = resultado['resposta']
            acao_executada = 'criar_lembrete'
            dados_criados = resultado.get('lembrete')
            
        elif any(palavra in comando for palavra in ['saldo', 'quanto', 'tenho', 'dinheiro']):
            # Consulta de saldo
            resultado = consultar_saldo_resumo(current_user.id)
            resposta = resultado['resposta']
            acao_executada = 'consultar_saldo'
            
        elif any(palavra in comando for palavra in ['gastos', 'gastei', 'categoria', 'onde']):
            # Consulta de gastos
            resultado = consultar_gastos_categoria(current_user.id)
            resposta = resultado['resposta']
            acao_executada = 'consultar_gastos'
            
        else:
            resposta = "Não entendi o comando financeiro. Você pode dizer coisas como 'gastei 50 reais com almoço' ou 'quanto tenho de saldo?'"
        
        return jsonify({
            'comando_original': comando,
            'resposta': resposta,
            'acao_executada': acao_executada,
            'dados_criados': dados_criados
        }), 200
        
    except Exception as e:
        return jsonify({'erro': f'Erro interno: {str(e)}'}), 500

def processar_comando_despesa(comando, user_id):
    """Processa comando de criação de despesa"""
    try:
        # Extrai valor do comando (busca por números)
        import re
        valores = re.findall(r'\d+(?:,\d+)?', comando)
        
        if not valores:
            return {'resposta': 'Não consegui identificar o valor da despesa. Tente dizer algo como "gastei 50 reais com almoço".'}
        
        valor = float(valores[0].replace(',', '.'))
        
        # Extrai descrição (remove valor e palavras comuns)
        descricao = comando
        for palavra in ['gastei', 'paguei', 'comprei', 'reais', 'real', 'com', str(valor)]:
            descricao = descricao.replace(palavra, '')
        descricao = descricao.strip()
        
        if not descricao:
            descricao = f'Despesa de R$ {valor:.2f}'
        
        # Busca categoria adequada ou usa padrão
        categoria = CategoriaFinanceira.query.filter_by(
            user_id=user_id,
            tipo='despesa',
            ativa=True
        ).first()
        
        if not categoria:
            # Cria categoria padrão
            categoria = CategoriaFinanceira(
                user_id=user_id,
                nome='Geral',
                tipo='despesa'
            )
            db.session.add(categoria)
            db.session.flush()
        
        # Cria transação
        transacao = TransacaoFinanceira(
            user_id=user_id,
            categoria_id=categoria.id,
            descricao=descricao.title(),
            valor=valor,
            tipo='despesa',
            data_transacao=date.today()
        )
        
        db.session.add(transacao)
        db.session.commit()
        
        return {
            'resposta': f'Registrei uma despesa de R$ {valor:.2f} com {descricao}.',
            'transacao': transacao.to_dict()
        }
        
    except Exception as e:
        return {'resposta': f'Erro ao processar despesa: {str(e)}'}

def processar_comando_receita(comando, user_id):
    """Processa comando de criação de receita"""
    try:
        import re
        valores = re.findall(r'\d+(?:,\d+)?', comando)
        
        if not valores:
            return {'resposta': 'Não consegui identificar o valor da receita. Tente dizer algo como "recebi 1000 reais de salário".'}
        
        valor = float(valores[0].replace(',', '.'))
        
        # Extrai descrição
        descricao = comando
        for palavra in ['recebi', 'ganhei', 'reais', 'real', 'de', str(valor)]:
            descricao = descricao.replace(palavra, '')
        descricao = descricao.strip()
        
        if not descricao:
            descricao = f'Receita de R$ {valor:.2f}'
        
        # Busca categoria de receita
        categoria = CategoriaFinanceira.query.filter_by(
            user_id=user_id,
            tipo='receita',
            ativa=True
        ).first()
        
        if not categoria:
            categoria = CategoriaFinanceira(
                user_id=user_id,
                nome='Salário',
                tipo='receita'
            )
            db.session.add(categoria)
            db.session.flush()
        
        # Cria transação
        transacao = TransacaoFinanceira(
            user_id=user_id,
            categoria_id=categoria.id,
            descricao=descricao.title(),
            valor=valor,
            tipo='receita',
            data_transacao=date.today()
        )
        
        db.session.add(transacao)
        db.session.commit()
        
        return {
            'resposta': f'Registrei uma receita de R$ {valor:.2f} - {descricao}.',
            'transacao': transacao.to_dict()
        }
        
    except Exception as e:
        return {'resposta': f'Erro ao processar receita: {str(e)}'}

def processar_comando_lembrete(comando, user_id):
    """Processa comando de criação de lembrete"""
    try:
        # Implementação básica - pode ser expandida
        return {
            'resposta': 'Funcionalidade de lembretes por voz em desenvolvimento. Use a interface para criar lembretes.',
            'lembrete': None
        }
        
    except Exception as e:
        return {'resposta': f'Erro ao processar lembrete: {str(e)}'}

def consultar_saldo_resumo(user_id):
    """Consulta saldo e resumo financeiro"""
    try:
        # Saldo das contas
        contas = ContaFinanceira.query.filter_by(
            user_id=user_id, 
            ativa=True, 
            incluir_no_total=True
        ).all()
        
        saldo_total = sum(float(conta.saldo_atual) for conta in contas)
        
        # Gastos do mês atual
        hoje = date.today()
        inicio_mes = date(hoje.year, hoje.month, 1)
        
        gastos_mes = TransacaoFinanceira.query.filter_by(user_id=user_id, tipo='despesa')\
                                           .filter(TransacaoFinanceira.data_transacao >= inicio_mes)\
                                           .all()
        
        total_gastos_mes = sum(float(t.valor) for t in gastos_mes)
        
        resposta = f'Seu saldo total é de R$ {saldo_total:.2f}. '
        resposta += f'Neste mês você gastou R$ {total_gastos_mes:.2f}.'
        
        return {'resposta': resposta}
        
    except Exception as e:
        return {'resposta': f'Erro ao consultar saldo: {str(e)}'}

def consultar_gastos_categoria(user_id):
    """Consulta gastos por categoria"""
    try:
        # Gastos dos últimos 30 dias
        data_limite = date.today() - timedelta(days=30)
        
        transacoes = TransacaoFinanceira.query.filter_by(user_id=user_id, tipo='despesa')\
                                            .filter(TransacaoFinanceira.data_transacao >= data_limite)\
                                            .all()
        
        gastos_categoria = {}
        for transacao in transacoes:
            if transacao.categoria:
                categoria = transacao.categoria.nome
                gastos_categoria[categoria] = gastos_categoria.get(categoria, 0) + float(transacao.valor)
        
        if not gastos_categoria:
            return {'resposta': 'Não encontrei gastos nos últimos 30 dias.'}
        
        # Top 3 categorias
        top_categorias = sorted(gastos_categoria.items(), key=lambda x: x[1], reverse=True)[:3]
        
        resposta = 'Seus maiores gastos dos últimos 30 dias: '
        for i, (categoria, valor) in enumerate(top_categorias):
            resposta += f'{categoria}: R$ {valor:.2f}'
            if i < len(top_categorias) - 1:
                resposta += ', '
        
        return {'resposta': resposta}
        
    except Exception as e:
        return {'resposta': f'Erro ao consultar gastos: {str(e)}'}

